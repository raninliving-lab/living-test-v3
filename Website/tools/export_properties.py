#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
DATABASE_DIR = ROOT / "DATA_BASE_ForRent"
MASTER_ZONE_PATH = ROOT / "DATA_BASE_ETC." / "chiangmai_master_zones.xlsx"
OUTPUT_PATH = ROOT / "Website" / "data" / "properties.json"
SCRIPT_OUTPUT_PATH = ROOT / "Website" / "data" / "properties.js"

WORKBOOKS = [
    {
        "file": "HomeForRent_ข้อมูลบ้านเช่า.xlsx",
        "asset_group": "HomeForRent",
        "category": "house",
        "category_label": "บ้านเช่า",
    },
    {
        "file": "CondoForRent_ข้อมูลคอนโด.xlsx",
        "asset_group": "CondoForRent",
        "category": "condo",
        "category_label": "คอนโดเช่า",
    },
    {
        "file": "PoolvillaForRent_ข้อมูลพูลวิลล่า.xlsx",
        "asset_group": "PoolvillaForRent",
        "category": "pool",
        "category_label": "พูลวิลล่า",
    },
]

FALLBACK_IMAGES = {
    "house": "https://images.unsplash.com/photo-1564013799919-ab600027ffc6?auto=format&fit=crop&w=1400&q=80",
    "condo": "https://images.unsplash.com/photo-1522708323590-d24dbb6b0267?auto=format&fit=crop&w=1400&q=80",
    "pool": "https://images.unsplash.com/photo-1613490493576-7fde63acd811?auto=format&fit=crop&w=1400&q=80",
}

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".jfif"}
VIDEO_EXTENSIONS = {".mp4", ".m4v", ".mov", ".webm"}
MAX_PROPERTY_IMAGES = 60
MAX_PROPERTY_VIDEOS = 12

ZONE_IMAGES = {
    "01": "../Zone/01 ในเมือง - คูเมือง - เมืองเก่า.jpg",
    "02": "../Zone/02 นิมมาน - ห้วยแก้ว - เมญ่า.jpg",
    "03": "../Zone/03 ช่างเคี่ยน - เจ็ดยอด.jpg",
    "04": "../Zone/04 หลัง มช - สุเทพ - โป่งน้อย.jpg",
    "05": "../Zone/05 รพ.ลานนา - ศูนย์ราชการ - โชตนา.jpg",
    "06": "../Zone/06 สันผีเสื้อ - รวมโชค - มีโชค.jpg",
    "07": "../Zone/07 เซ็นเฟส - ฟ้าฮ่าม - พายัพ.jpg",
    "08": "../Zone/08 ช้างคลาน - ไนท์บาซาร์.jpg",
    "09": "../Zone/09 พรอมเมนาดา - ดอนจั่น - ท่าศาลา.jpg",
    "10": "../Zone/10 แม่เหียะ - สนามบิน - ราชพฤกษ์.jpg",
    "11": "../Zone/11 หางดง - กาดฝรั่ง - บ้านถวาย.jpg",
    "12": "../Zone/12 แม่ริม.jpg",
    "13": "../Zone/13 สันทราย - แม่โจ้.jpg",
    "14": "../Zone/14 ดอยสะเก็ด - สันปูเลย.jpg",
    "15": "../Zone/15 สันกำแพง - บ่อสร้าง.jpg",
    "16": "../Zone/16 หนองหอย - สารภี - ยางเนิ้ง.jpg",
}

TRACKING_LABELS = {
    "available": "ว่างตอนนี้",
    "soon": "กำลังจะว่าง",
    "follow_up": "ควรติดตาม",
    "occupied": "ไม่ว่าง",
    "unknown": "ไม่ระบุวันว่าง",
}

PUBLIC_MAP_QUERIES_BY_ZONE = {
    "01": "ตำบลศรีภูมิ เชียงใหม่",
    "02": "ตำบลสุเทพ เชียงใหม่",
    "03": "ตำบลช้างเผือก เชียงใหม่",
    "04": "ตำบลสุเทพ เชียงใหม่",
    "05": "ตำบลช้างเผือก เชียงใหม่",
    "06": "ตำบลสันผีเสื้อ เชียงใหม่",
    "07": "ตำบลฟ้าฮ่าม เชียงใหม่",
    "08": "ตำบลช้างคลาน เชียงใหม่",
    "09": "ตำบลท่าศาลา เชียงใหม่",
    "10": "ตำบลแม่เหียะ เชียงใหม่",
    "11": "ตำบลหางดง เชียงใหม่",
    "12": "ตำบลแม่ริม เชียงใหม่",
    "13": "ตำบลแม่โจ้ เชียงใหม่",
    "14": "ตำบลสันปูเลย เชียงใหม่",
    "15": "ตำบลสันกำแพง เชียงใหม่",
    "16": "ตำบลหนองหอย เชียงใหม่",
}

PUBLIC_MAP_QUERIES_BY_SUBZONE = {
    "สันผีเสื้อ": "ตำบลสันผีเสื้อ เชียงใหม่",
    "เซ็นเฟส": "ตำบลฟ้าฮ่าม เชียงใหม่",
    "พายัพ": "ตำบลฟ้าฮ่าม เชียงใหม่",
    "พรอม": "ตำบลท่าศาลา เชียงใหม่",
    "ดอนจั่น": "ตำบลท่าศาลา เชียงใหม่",
    "แม่เหียะ": "ตำบลแม่เหียะ เชียงใหม่",
    "หางดง": "ตำบลหางดง เชียงใหม่",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def get_value(row: tuple[Any, ...], headers: dict[str, int], *names: str) -> Any:
    for name in names:
        index = headers.get(name)
        if index is not None and index < len(row):
            value = row[index]
            if value not in (None, ""):
                return value
    return None


def is_url(value: str) -> bool:
    return bool(re.match(r"^https?://", value.strip(), re.IGNORECASE))


def public_map_query(row: tuple[Any, ...], headers: dict[str, int], zone_code: str, zone_label: str) -> str:
    public_location = clean_text(get_value(row, headers, "Location แสดงบนเว็บ URL", "ตำบลมาตรฐาน"))
    if public_location and not is_url(public_location):
        return public_location

    subzone = re.sub(r"^\d+", "", clean_text(get_value(row, headers, "SubZone ใหม่"))).strip()
    for keyword, query in PUBLIC_MAP_QUERIES_BY_SUBZONE.items():
        if keyword in subzone:
            return query

    return PUBLIC_MAP_QUERIES_BY_ZONE.get(zone_code) or f"{zone_label} เชียงใหม่"


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group(0)) if match else None


def number_or_none(value: Any) -> int | float | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return int(parsed) if parsed.is_integer() else parsed


def bool_or_none(value: Any) -> bool | None:
    text = clean_text(value).lower()
    if not text:
        return None
    negative_tokens = ("ไม่รับ", "ไม่มี", "ไม่รวม", "ไม่", "no", "false", "x", "✕")
    positive_tokens = ("รับ", "มี", "รวม", "ใช่", "yes", "true", "✓", "check")
    if any(token in text for token in negative_tokens) or text in {"0", "-", "—"}:
        return False
    if any(token in text for token in positive_tokens) or text == "1":
        return True
    return None


def detail_number(row: tuple[Any, ...], headers: dict[str, int], *names: str) -> int | float | None:
    return number_or_none(get_value(row, headers, *names))


def detail_bool(row: tuple[Any, ...], headers: dict[str, int], *names: str) -> bool | None:
    return bool_or_none(get_value(row, headers, *names))


def lease_details(row: tuple[Any, ...], headers: dict[str, int]) -> dict[str, Any]:
    return {
        "contractTermMonths": detail_number(row, headers, "ระยะเวลาตามสัญญา (เดือน)", "ระยะสัญญาหลัก (เดือน)", "ระยะเวลาตามสัญญา"),
        "securityDepositMonths": detail_number(row, headers, "เงินประกัน (เดือน)", "ประกัน (เดือน)", "เงินประกัน"),
        "advanceRentMonths": detail_number(row, headers, "ค่าเช่าล่วงหน้า (เดือน)", "ล่วงหน้า (เดือน)", "ค่าเช่าล่วงหน้า"),
        "moveInTotalMonths": detail_number(row, headers, "ยอดชำระวันเข้าอยู่ (เดือน)", "ยอดชำระเข้าอยู่ (เดือน)", "ยอดชำระวันเข้าอยู่"),
        "shortLeases": {
            "6-11": {
                "accepted": detail_bool(row, headers, "รับสัญญา 6-11 เดือน", "6-11 เดือน รับหรือไม่"),
                "premiumPercent": detail_number(row, headers, "เพิ่มราคา 6-11 เดือน (%)", "6-11 เดือน เพิ่ม (%)", "6-11 เดือน เพิ่มจากราคาฐาน (%)"),
            },
            "1-5": {
                "accepted": detail_bool(row, headers, "รับสัญญา 1-5 เดือน", "1-5 เดือน รับหรือไม่"),
                "premiumPercent": detail_number(row, headers, "เพิ่มราคา 1-5 เดือน (%)", "1-5 เดือน เพิ่ม (%)", "1-5 เดือน เพิ่มจากราคาฐาน (%)"),
            },
        },
        "includedInRent": {
            "commonFee": detail_bool(row, headers, "รวมค่าส่วนกลาง", "ค่าส่วนกลางรวม"),
            "internet": detail_bool(row, headers, "รวมอินเทอร์เน็ต", "อินเทอร์เน็ตรวม"),
            "poolCare": detail_bool(row, headers, "รวมดูแลสระ", "ดูแลสระรวม", "บริการดูแลสระว่ายน้ำรวม"),
            "gardenCare": detail_bool(row, headers, "รวมดูแลสวน", "ดูแลสวนรวม", "บริการดูแลสวนรวม"),
        },
        "tenantResponsible": {
            "water": detail_bool(row, headers, "ผู้เช่าจ่ายค่าน้ำ", "ค่าน้ำผู้เช่ารับผิดชอบ"),
            "electricity": detail_bool(row, headers, "ผู้เช่าจ่ายค่าไฟ", "ค่าไฟผู้เช่ารับผิดชอบ"),
            "internet": detail_bool(row, headers, "ผู้เช่าจ่ายอินเทอร์เน็ต", "อินเทอร์เน็ตผู้เช่ารับผิดชอบ"),
            "otherUsage": detail_bool(row, headers, "ผู้เช่าจ่ายค่าใช้จ่ายอื่น", "ค่าบริการอื่นตามการใช้งาน"),
        },
        "appliances": {
            "airConditioner": detail_number(row, headers, "จำนวนแอร์", "จำนวนเครื่องปรับอากาศ"),
            "television": detail_number(row, headers, "จำนวนทีวี", "จำนวนโทรทัศน์"),
            "refrigerator": detail_number(row, headers, "จำนวนตู้เย็น"),
            "washingMachine": detail_number(row, headers, "จำนวนเครื่องซักผ้า"),
            "waterHeater": detail_number(row, headers, "จำนวนเครื่องทำน้ำอุ่น"),
            "stoveHood": detail_number(row, headers, "จำนวนเตาและเครื่องดูดควัน", "จำนวนชุดเตาและเครื่องดูดควัน"),
        },
        "furniture": {
            "bedMattress": detail_number(row, headers, "จำนวนเตียงและที่นอน", "จำนวนเตียง"),
            "wardrobe": detail_number(row, headers, "จำนวนตู้เสื้อผ้า"),
            "sofa": detail_number(row, headers, "จำนวนโซฟา"),
            "diningTable": detail_number(row, headers, "จำนวนโต๊ะรับประทานอาหาร"),
        },
        "notes": clean_text(get_value(row, headers, "หมายเหตุอื่นๆ", "หมายเหตุอื่น ๆ", "หมายเหตุเฟอร์นิเจอร์")),
    }


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            if parsed.year > 2400:
                parsed = parsed.replace(year=parsed.year - 543)
            return parsed
        except ValueError:
            pass
    return None


def display_date(value: date | None) -> str:
    if not value:
        return ""
    return value.strftime("%d/%m/%Y")


def normalize_zone_label(value: str) -> str:
    return re.sub(r"\s*/\s*", " / ", value).strip()


def load_master_zones() -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    workbook = openpyxl.load_workbook(MASTER_ZONE_PATH, read_only=True, data_only=True)
    sheet = workbook["Master Zones"]
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {clean_text(value): index for index, value in enumerate(header_values) if clean_text(value)}
    zones: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = clean_text(row[headers["โซน"]]).zfill(2)
        title = normalize_zone_label(clean_text(row[headers["ชื่อโซน"]]))
        areas = clean_text(row[headers["พื้นที่ที่รวมอยู่"]])
        keyword = clean_text(row[headers["Keyword ขาย"]])
        reason = clean_text(row[headers["ปัจจัยหลักที่ทำให้เลือก"]])
        zones.append(
            {
                "code": code,
                "no": code,
                "title": title,
                "label": title,
                "zoneLabel": title,
                "areas": areas,
                "keyword": keyword,
                "eyebrow": keyword,
                "desc": reason,
                "assetZones": [code],
                "image": ZONE_IMAGES.get(code, ""),
                "supermarket": clean_text(row[headers["ห้าง / Supermarket"]]),
                "hospital": clean_text(row[headers["โรงพยาบาล"]]),
                "school": clean_text(row[headers["International / Bilingual Schools"]]),
                "university": clean_text(row[headers["University"]]),
                "market": clean_text(row[headers["ตลาด"]]),
                "petService": clean_text(row[headers["Pet Service"]]),
                "airport": clean_text(row[headers["Airport"]]),
            }
        )

    workbook.close()
    return zones, {zone["code"]: zone for zone in zones}


def zone_parts(zone_value: Any, zones_by_code: dict[str, dict[str, Any]]) -> tuple[str, str]:
    zone = clean_text(zone_value)
    match = re.match(r"Z(\d+)_?(.*)", zone)
    if not match:
        return "", zone
    code = match.group(1).zfill(2)
    master_zone = zones_by_code.get(code)
    if master_zone:
        return code, master_zone["label"]
    label = match.group(2).replace("_", " ").strip() or zone
    return code, label


def public_path(path: Path) -> str:
    rel = path.relative_to(ROOT).as_posix()
    return f"../{rel}"


def natural_sort_key(path: Path) -> list[Any]:
    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", path.name)
    ]


def media_files(directory: Path, extensions: set[str]) -> list[Path]:
    if not directory.exists():
        return []
    return sorted(
        (
            path
            for path in directory.glob("*")
            if path.is_file() and path.suffix.lower() in extensions
        ),
        key=natural_sort_key,
    )


def collect_property_media(folder: Path) -> dict[str, list[str]]:
    final_dir = folder / "03_Final" / "01_Photos_Final"
    photos = media_files(final_dir, IMAGE_EXTENSIONS)
    covers = [path for path in photos if "cover" in path.name.lower()]
    gallery = (covers + [path for path in photos if path not in covers])[:MAX_PROPERTY_IMAGES]

    video_dir = folder / "03_Final" / "02_Videos_Final"
    videos = media_files(video_dir, VIDEO_EXTENSIONS)[:MAX_PROPERTY_VIDEOS]

    return {
        "images": [public_path(path) for path in gallery],
        "videos": [public_path(path) for path in videos],
    }


def find_post_url(folder: Path) -> str:
    post_dir = folder / "03_Final" / "03_Post_ประกาศ"
    txt_files = sorted(post_dir.glob("*.txt"))
    docx_files = sorted(post_dir.glob("*.docx"))
    candidates = txt_files or docx_files
    return public_path(candidates[-1]) if candidates else ""


def folder_has_images(folder: Path) -> bool:
    return bool(media_files(folder / "03_Final" / "01_Photos_Final", IMAGE_EXTENSIONS))


def find_property_folder(asset_group: str, property_id: str, folder_value: str) -> Path | None:
    candidates: list[Path] = []
    folder = ROOT / folder_value if folder_value else None

    if folder:
        candidates.append(folder)
        if re.fullmatch(r"\d+", folder.name) and re.match(rf"^{re.escape(property_id)}_", folder.parent.name):
            candidates.append(folder.parent)

    asset_root = ROOT / "asset" / asset_group
    fallback = next(asset_root.glob(f"*/{property_id}_*"), None) if asset_root.exists() else None
    if fallback:
        candidates.append(fallback)

    unique_candidates = list(dict.fromkeys(candidates))
    for candidate in unique_candidates:
        if candidate.exists() and folder_has_images(candidate):
            return candidate
    return next((candidate for candidate in unique_candidates if candidate.exists()), None)


def title_from_folder(folder: Path, fallback: str) -> str:
    name = folder.name if folder else ""
    if re.fullmatch(r"\d+", name) and re.match(r"^Z\d{2}[A-Z]\d{3}_", folder.parent.name):
        name = f"{folder.parent.name}/{name}"
    if "_" in name:
        name = name.split("_", 1)[1]
    name = name.replace("_", " ").strip()

    fallback_text = clean_text(fallback)
    if fallback_text and "/" in fallback_text and "/" not in name:
        fallback_base, fallback_suffix = [part.strip() for part in fallback_text.rsplit("/", 1)]
        fallback_base_norm = re.sub(r"\s+", "", fallback_base).lower()
        name_norm = re.sub(r"\s+", "", name).lower()
        if fallback_suffix and fallback_base_norm and name_norm.endswith(fallback_base_norm):
            name = f"{name}/{fallback_suffix}"

    return name or fallback


def classify_tracking(status_value: Any, contract_end: date | None, today: date) -> dict[str, str]:
    status = clean_text(status_value)
    if "ทรัพว่าง" in status or status == "ว่าง":
        group = "available"
    elif "ติดต่อไม่ได้" in status or "ไม่ระบุ" in status:
        group = "unknown"
    elif contract_end:
        days = (contract_end - today).days
        if 0 <= days <= 90:
            group = "soon"
        elif days < 0 or days <= 180:
            group = "follow_up"
        else:
            group = "occupied"
    elif "เช่าแล้ว" in status:
        group = "occupied"
    else:
        group = "unknown"
    return {"group": group, "label": TRACKING_LABELS[group], "rawStatus": status}


def area_fields(row: tuple[Any, ...], headers: dict[str, int], category: str) -> tuple[int | float | None, str]:
    if category == "condo":
        area = number_or_none(get_value(row, headers, "ขนาดพื้นที่", "ขนาดพื้นที่ (ตร.ม.)"))
        return area, f"{area:g} ตร.ม." if isinstance(area, (int, float)) else ""
    area = number_or_none(get_value(row, headers, "ขนาดพื้นที่ (ตร.ว.)", "ขนาดพื้นที่", "พื้นที่"))
    return area, f"{area:g} ตร.ว." if isinstance(area, (int, float)) else ""


def should_skip(status_value: Any) -> bool:
    status = clean_text(status_value)
    return any(word in status for word in ("ยกเลิก", "ลบ", "ไม่ใช้"))


def export_properties() -> dict[str, Any]:
    today = date.today()
    properties: list[dict[str, Any]] = []
    zones, zones_by_code = load_master_zones()

    for config in WORKBOOKS:
        workbook_path = DATABASE_DIR / config["file"]
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook["MASTER_DATA"]
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {clean_text(value): index for index, value in enumerate(header_values) if clean_text(value)}

        for excel_row, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            property_id = clean_text(get_value(row, headers, "รหัสทรัพย์"))
            if not property_id:
                continue

            status = get_value(row, headers, "สถานะทรัพย์")
            if should_skip(status):
                continue

            folder_value = clean_text(get_value(row, headers, "Path โฟลเดอร์"))
            folder = find_property_folder(config["asset_group"], property_id, folder_value)

            zone_code, zone_label = zone_parts(get_value(row, headers, "โซน"), zones_by_code)
            contract_end = parse_date(
                get_value(row, headers, "วันครบกำหนด", "วันสิ้นสุดสัญญา", "วันหมดสัญญา")
            )
            media = collect_property_media(folder) if folder else {"images": [], "videos": []}
            image = media["images"][0] if media["images"] else FALLBACK_IMAGES[config["category"]]
            image_source = "local" if media["images"] else "fallback"
            area_value, area_label = area_fields(row, headers, config["category"])
            display_location = clean_text(
                get_value(
                    row,
                    headers,
                    "Location แสดงบนเว็บ",
                    "ที่ตั้ง",
                    "ตำบลมาตรฐาน",
                    "SubZone ใหม่",
                    "โซน",
                )
            )
            title = title_from_folder(folder, display_location or config["category_label"]) if folder else (
                display_location or property_id
            )

            properties.append(
                {
                    "id": property_id,
                    "category": config["category"],
                    "categoryLabel": config["category_label"],
                    "title": title,
                    "zoneCode": zone_code,
                    "zoneLabel": zone_label,
                    "displayLocation": display_location or zone_label,
                    "rent": number_or_none(get_value(row, headers, "ค่าเช่า/เดือน")),
                    "bedrooms": number_or_none(get_value(row, headers, "ห้องนอน")),
                    "bathrooms": number_or_none(get_value(row, headers, "ห้องน้ำ")),
                    "areaSqWa": area_value,
                    "areaLabel": area_label,
                    "contractEndDisplay": display_date(contract_end),
                    "tracking": classify_tracking(status, contract_end, today),
                    "leaseDetails": lease_details(row, headers),
                    "image": image,
                    "imageSource": image_source,
                    "media": media,
                    "postUrl": find_post_url(folder) if folder else "",
                    "mapUrl": public_map_query(row, headers, zone_code, zone_label),
                    "photoDriveUrl": clean_text(get_value(row, headers, "Link Google Drive รูปภาพ")),
                    "source": {
                        "workbook": config["file"],
                        "sheet": "MASTER_DATA",
                        "row": excel_row,
                        "folder": folder.relative_to(ROOT).as_posix() if folder else "",
                    },
                }
            )

        workbook.close()

    sort_order = {"available": 0, "soon": 1, "follow_up": 2, "unknown": 3, "occupied": 4}
    properties.sort(
        key=lambda item: (
            item["category"],
            sort_order.get(item["tracking"]["group"], 9),
            int(item["zoneCode"] or 999),
            item["id"],
        )
    )

    counts: dict[str, int] = {}
    for item in properties:
        counts[item["category"]] = counts.get(item["category"], 0) + 1

    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "today": today.isoformat(),
        "source": "DATA_BASE_ForRent Excel export",
        "counts": counts,
        "zones": zones,
        "properties": properties,
    }


def main() -> None:
    data = export_properties()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    json_text = json.dumps(data, ensure_ascii=False, indent=2)
    OUTPUT_PATH.write_text(json_text, encoding="utf-8")
    SCRIPT_OUTPUT_PATH.write_text(
        "window.PRANA_PROPERTIES_DATA = " + json_text + ";\n",
        encoding="utf-8",
    )
    print(
        f"Wrote {OUTPUT_PATH.relative_to(ROOT)} with "
        f"{len(data['properties']):,} properties"
    )


if __name__ == "__main__":
    main()
